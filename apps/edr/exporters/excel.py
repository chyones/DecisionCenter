"""Excel (.xlsx) exporter — for evidence tables, financial summaries, audit data."""

from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_BOLD_FONT = Font(bold=True)


def _header_row(ws, values: list[str], row: int = 1) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _set_col_widths(ws, widths: list[int]) -> None:
    letters = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    for i, w in enumerate(widths):
        if i < len(letters):
            ws.column_dimensions[letters[i]].width = w


def to_excel(report: dict) -> bytes:
    """Generate a multi-sheet Excel workbook from the canonical JSON report dict."""
    wb = Workbook()

    request_id = report.get("request_id", "N/A")
    project_code = report.get("project_code") or "N/A"
    query = report.get("query") or report.get("question", "N/A")
    language = report.get("language", "en")
    qg_status = report.get("quality_gate_status", "not_run")
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _header_row(ws, ["Field", "Value"])
    for r, (key, val) in enumerate([
        ("Request ID", request_id),
        ("Project", project_code),
        ("Language", language),
        ("Quality Gate", qg_status),
        ("Generated", generated_at),
        ("Query", query),
    ], start=2):
        ws.cell(row=r, column=1, value=key).font = _BOLD_FONT
        ws.cell(row=r, column=2, value=val)

    ws.cell(row=9, column=1, value="Executive Summary").font = _BOLD_FONT
    summary = report.get("executive_summary", [])
    row_i = 10
    if isinstance(summary, list):
        for item in summary:
            if isinstance(item, dict):
                ws.cell(row=row_i, column=1, value=item.get("claim", ""))
                ws.cell(row=row_i, column=2, value=item.get("confidence", "medium"))
                row_i += 1
    _set_col_widths(ws, [25, 80])

    # ── Sheet 2: Financial Snapshot ───────────────────────────────────────────
    ws_fin = wb.create_sheet("Financial Snapshot")
    _header_row(ws_fin, ["Item", "Value", "Currency", "Evidence ID", "Status"])
    fs = report.get("financial_snapshot") or {}
    for r, (label, node) in enumerate([
        ("Budget", (fs.get("budget") or {}) if isinstance(fs, dict) else {}),
        ("Actual Cost", (fs.get("actual_cost") or {}) if isinstance(fs, dict) else {}),
    ], start=2):
        ws_fin.cell(row=r, column=1, value=label)
        if isinstance(node, dict):
            ws_fin.cell(row=r, column=2, value=node.get("value"))
            ws_fin.cell(row=r, column=3, value=node.get("currency", "AED"))
            ws_fin.cell(row=r, column=4, value=node.get("evidence_id") or "—")
            ws_fin.cell(row=r, column=5, value=node.get("status", "not_available"))
        else:
            ws_fin.cell(row=r, column=2, value="Not available")

    if isinstance(fs, dict):
        variance = fs.get("variance") or {}
        if isinstance(variance, dict):
            ws_fin.cell(row=4, column=1, value="Variance")
            ws_fin.cell(row=4, column=2, value=variance.get("value"))
            ws_fin.cell(row=4, column=3, value=variance.get("currency", "AED"))
            ws_fin.cell(row=4, column=4, value=variance.get("formula") or "—")
            ws_fin.cell(row=4, column=5, value="calculated")
    _set_col_widths(ws_fin, [15, 18, 10, 20, 16])

    # ── Sheet 3: Sources ──────────────────────────────────────────────────────
    ws_src = wb.create_sheet("Sources")
    _header_row(ws_src, ["Source ID", "Type", "Title", "Reference", "Date", "Confidence", "Used In"])
    for r, src in enumerate(report.get("sources", []), start=2):
        if isinstance(src, dict):
            ws_src.cell(row=r, column=1, value=src.get("source_id", "—"))
            ws_src.cell(row=r, column=2, value=src.get("source_type", "—"))
            ws_src.cell(row=r, column=3, value=src.get("title", "—"))
            ws_src.cell(row=r, column=4, value=src.get("reference", "—"))
            ws_src.cell(row=r, column=5, value=src.get("date") or "—")
            ws_src.cell(row=r, column=6, value=src.get("confidence", "—"))
            ws_src.cell(row=r, column=7, value=", ".join(src.get("used_in", [])))
    _set_col_widths(ws_src, [12, 12, 30, 40, 12, 12, 30])

    # ── Sheet 4: Conflicts ────────────────────────────────────────────────────
    ws_conf = wb.create_sheet("Conflicts")
    _header_row(ws_conf, ["Conflict Type", "Description", "Source A", "Source B", "Conf. A", "Conf. B"])
    for r, c in enumerate(report.get("conflicts", []), start=2):
        if isinstance(c, dict):
            ws_conf.cell(row=r, column=1, value=c.get("conflict_type", "—"))
            ws_conf.cell(row=r, column=2, value=c.get("description", "—"))
            ws_conf.cell(row=r, column=3, value=c.get("source_a_ref", "—"))
            ws_conf.cell(row=r, column=4, value=c.get("source_b_ref", "—"))
            ws_conf.cell(row=r, column=5, value=c.get("confidence_a", "—"))
            ws_conf.cell(row=r, column=6, value=c.get("confidence_b", "—"))
    _set_col_widths(ws_conf, [20, 50, 20, 20, 10, 10])

    # ── Sheet 5: Missing Data ─────────────────────────────────────────────────
    ws_miss = wb.create_sheet("Missing Data")
    _header_row(ws_miss, ["#", "Missing Item"])
    for r, item in enumerate(report.get("missing_data", []), start=2):
        ws_miss.cell(row=r, column=1, value=r - 1)
        ws_miss.cell(row=r, column=2, value=str(item))
    _set_col_widths(ws_miss, [5, 80])

    # ── Sheet 6: Audit Data (all findings) ────────────────────────────────────
    ws_audit = wb.create_sheet("Audit Data")
    _header_row(ws_audit, ["Section", "Finding", "Confidence", "Evidence IDs"])
    audit_row = 2
    for section_name, key in [
        ("Key Findings", "key_findings"),
        ("Root Causes", "root_causes"),
        ("Delay Analysis", "delay_analysis"),
        ("Contractual Implications", "contractual_implications"),
        ("Recommended Actions", "recommended_actions"),
    ]:
        for item in report.get(key, []):
            if isinstance(item, dict):
                ws_audit.cell(row=audit_row, column=1, value=section_name)
                ws_audit.cell(row=audit_row, column=2, value=item.get("text", ""))
                ws_audit.cell(row=audit_row, column=3, value=item.get("confidence", "medium"))
                ws_audit.cell(
                    row=audit_row, column=4, value=", ".join(item.get("evidence_ids", []))
                )
                audit_row += 1
    _set_col_widths(ws_audit, [25, 60, 12, 30])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
